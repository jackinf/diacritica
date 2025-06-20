import tkinter as tk
from tkinter import ttk, messagebox
from tkinterdnd2 import TkinterDnD, DND_FILES
import openpyxl
from openpyxl.utils import get_column_letter
import os
import sys
from pathlib import Path

class DiacriticsRemover:
    def __init__(self):
        # Create the diacritics mapping dictionary
        self.diacritics_map = {
            'À': 'A', 'Á': 'A', 'Â': 'A', 'Ã': 'A', 'Ä': 'A', 'Å': 'A', 'Æ': 'AE',
            'Ç': 'C', 'È': 'E', 'É': 'E', 'Ê': 'E', 'Ë': 'E',
            'Ì': 'I', 'Í': 'I', 'Î': 'I', 'Ï': 'I', 'Ð': 'D', 'Ñ': 'N',
            'Ò': 'O', 'Ó': 'O', 'Ô': 'O', 'Õ': 'O', 'Ö': 'O', 'Ø': 'O',
            'Ù': 'U', 'Ú': 'U', 'Û': 'U', 'Ü': 'U', 'Ý': 'Y', 'Þ': 'TH',
            'ß': 'ss', 'à': 'a', 'á': 'a', 'â': 'a', 'ã': 'a', 'ä': 'a',
            'å': 'a', 'æ': 'ae', 'ç': 'c', 'è': 'e', 'é': 'e', 'ê': 'e',
            'ë': 'e', 'ì': 'i', 'í': 'i', 'î': 'i', 'ï': 'i', 'ð': 'd',
            'ñ': 'n', 'ò': 'o', 'ó': 'o', 'ô': 'o', 'õ': 'o', 'ö': 'o',
            'ø': 'o', 'ù': 'u', 'ú': 'u', 'û': 'u', 'ü': 'u', 'ý': 'y',
            'þ': 'th', 'ÿ': 'y'
        }
        
    def remove_diacritics(self, text):
        """Remove diacritics from a string"""
        if not isinstance(text, str):
            return text
        
        result = []
        i = 0
        while i < len(text):
            char = text[i]
            if char in self.diacritics_map:
                result.append(self.diacritics_map[char])
            else:
                result.append(char)
            i += 1
        
        return ''.join(result)
    
    def process_excel_file(self, file_path):
        """Process an Excel file and remove diacritics from all cells"""
        try:
            # Load the workbook
            wb = openpyxl.load_workbook(file_path)
            
            # Process each worksheet
            for ws in wb.worksheets:
                # Get the dimensions of the used range
                for row in ws.iter_rows():
                    for cell in row:
                        # Skip cells with formulas
                        if cell.value is not None and not hasattr(cell, 'formula') or (hasattr(cell, 'formula') and cell.formula is None):
                            # Process the cell value
                            cell.value = self.remove_diacritics(str(cell.value))
            
            # Create output filename
            input_path = Path(file_path)
            output_path = input_path.parent / f"{input_path.stem}_fixed{input_path.suffix}"
            
            # Save the modified workbook
            wb.save(output_path)
            wb.close()
            
            return True, f"File processed successfully!\nSaved as: {output_path.name}"
            
        except Exception as e:
            return False, f"Error processing file: {str(e)}"

class DiacriticsRemoverGUI:
    def __init__(self):
        self.remover = DiacriticsRemover()
        self.current_file = None
        
        # Create main window with drag and drop support
        self.root = TkinterDnD.Tk()
        self.root.title("Excel Diacritics Remover")
        self.root.geometry("500x400")
        self.root.resizable(False, False)
        
        # Set window style
        self.root.configure(bg='#f0f0f0')
        
        # Create main frame
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title label
        title_label = ttk.Label(main_frame, text="Excel Diacritics Remover", 
                               font=('Arial', 16, 'bold'))
        title_label.pack(pady=(0, 20))
        
        # Create drop zone
        self.drop_frame = tk.Frame(main_frame, bg='white', relief=tk.GROOVE, 
                                  borderwidth=2, height=200)
        self.drop_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        
        # Drop zone label
        self.drop_label = tk.Label(self.drop_frame, 
                                  text="Drag and drop Excel file here\n(.xlsx, .xlsm, .xls)",
                                  bg='white', fg='#666666', font=('Arial', 12))
        self.drop_label.place(relx=0.5, rely=0.5, anchor=tk.CENTER)
        
        # File name label
        self.file_label = ttk.Label(main_frame, text="No file selected", 
                                   font=('Arial', 10))
        self.file_label.pack(pady=(0, 10))
        
        # Fix button
        self.fix_button = ttk.Button(main_frame, text="Fix", 
                                    command=self.process_file, 
                                    state=tk.DISABLED,
                                    style='Accent.TButton')
        self.fix_button.pack(pady=(0, 10))
        
        # Configure drag and drop
        self.drop_frame.drop_target_register(DND_FILES)
        self.drop_frame.dnd_bind('<<Drop>>', self.drop_file)
        self.drop_frame.dnd_bind('<<DragEnter>>', self.drag_enter)
        self.drop_frame.dnd_bind('<<DragLeave>>', self.drag_leave)
        
        # Configure button style
        style = ttk.Style()
        style.configure('Accent.TButton', font=('Arial', 12, 'bold'))
        
    def drag_enter(self, event):
        """Handle drag enter event"""
        self.drop_frame.configure(bg='#e6f3ff')
        self.drop_label.configure(bg='#e6f3ff')
        
    def drag_leave(self, event):
        """Handle drag leave event"""
        self.drop_frame.configure(bg='white')
        self.drop_label.configure(bg='white')
        
    def drop_file(self, event):
        """Handle file drop event"""
        self.drop_frame.configure(bg='white')
        self.drop_label.configure(bg='white')
        
        # Get the dropped file path
        file_path = event.data
        
        # Clean up the file path (remove curly braces if present)
        if file_path.startswith('{') and file_path.endswith('}'):
            file_path = file_path[1:-1]
        
        # Check if it's an Excel file
        valid_extensions = ('.xlsx', '.xlsm', '.xls')
        if file_path.lower().endswith(valid_extensions):
            self.current_file = file_path
            file_name = os.path.basename(file_path)
            self.file_label.configure(text=f"Selected: {file_name}")
            self.fix_button.configure(state=tk.NORMAL)
            self.drop_label.configure(text=f"✓ {file_name}")
        else:
            messagebox.showerror("Invalid File", 
                               "Please drop a valid Excel file (.xlsx, .xlsm, .xls)")
            
    def process_file(self):
        """Process the selected Excel file"""
        if not self.current_file:
            return
        
        # Disable button during processing
        self.fix_button.configure(state=tk.DISABLED, text="Processing...")
        self.root.update()
        
        # Process the file
        success, message = self.remover.process_excel_file(self.current_file)
        
        # Re-enable button
        self.fix_button.configure(state=tk.NORMAL, text="Fix")
        
        # Show result
        if success:
            messagebox.showinfo("Done", message)
            # Reset the interface
            self.current_file = None
            self.file_label.configure(text="No file selected")
            self.drop_label.configure(text="Drag and drop Excel file here\n(.xlsx, .xlsm, .xls)")
            self.fix_button.configure(state=tk.DISABLED)
        else:
            messagebox.showerror("Error", message)
    
    def run(self):
        """Start the GUI application"""
        self.root.mainloop()

if __name__ == "__main__":
    app = DiacriticsRemoverGUI()
    app.run()